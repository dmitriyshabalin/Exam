import os 
import sys
from InterfaceBD import *
from PyQt5 import QtCore, QtGui, QtWidgets
import sqlite3
from PyQt5.QtWidgets import QTableWidgetItem as QWT
from PyQt5.QtWidgets import QMessageBox
from PyQt5 import Qt
import openpyxl

class MyWin(QtWidgets.QMainWindow):
    def __init__(self, parent=None):
        QtWidgets.QWidget.__init__(self, parent)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        

        self.ui.pushButton.clicked.connect(self.Podkl)

        self.ui.pushButton_2.clicked.connect(self.Update) # Проверка выбранной таблицы

        # Изменение записей таблицы
        self.ui.pushButton_6.clicked.connect(self.upd1)
        self.ui.pushButton_8.clicked.connect(self.upd2)
        self.ui.pushButton_10.clicked.connect(self.upd3)
        self.ui.pushButton_12.clicked.connect(self.upd4)

        # Добавление записей
        self.ui.pushButton_14.clicked.connect(self.add1)
        self.ui.pushButton_15.clicked.connect(self.add2)
        self.ui.pushButton_16.clicked.connect(self.add3)
        self.ui.pushButton_17.clicked.connect(self.add4)

        self.ui.pushButton_3.clicked.connect(self.del1) #Удаление записей
        

        # Кнопки отмены
        self.ui.pushButton_7.clicked.connect(self.Cancel)
        self.ui.pushButton_9.clicked.connect(self.Cancel)
        self.ui.pushButton_11.clicked.connect(self.Cancel)
        self.ui.pushButton_13.clicked.connect(self.Cancel)
        
        self.ui.pushButton_5.clicked.connect(self.close) # Выход

        self.ui.statusbar.showMessage("База данных 'registr'" , 5000)

        self.ui.action_2.triggered.connect(self.Admin) # Переключиться в режим администратора
        self.ui.action_4.triggered.connect(self.Pol) # Переключиться в режим пользователя

        self.ui.action_6.triggered.connect(self.Prog) # О программе
        
        self.ui.actionWord.triggered.connect(lambda: self.Seal(1))# Сохранение
        self.ui.action_3.triggered.connect(lambda: self.Seal(2))
        self.ui.actionExcel.triggered.connect(lambda: self.Seal(3))
        
        self.ui.action.triggered.connect(self.Instruk) # Инструкция пользователя

        #! По умолчанию включен режим пользователя
        self.ui.pushButton_2.setEnabled(False)
        self.ui.pushButton_3.setEnabled(False)

        self.ui.pushButton_2.setStyleSheet('''QPushButton {
                                            background-color: ##424242;
                                            color: grey;
                                            }''')

        self.ui.pushButton_3.setStyleSheet('''QPushButton {
                                            background-color: ##424242;
                                            color: grey;
                                            }''')

        self.flacon = 0 # Переменная для проверки текущего режима

        self.ui.stackedWidget_2.setCurrentIndex(1)

        # Проверка наличия установленного пароля
        self.flag_admin = open('Pass.txt','r')
        for i in self.flag_admin:
            self.ing = i
            
        
        self.ui.tabWidget.currentChanged.connect(self.pri)

        self.Pods() # Функция подсказок

        #Валидаторы
        rx = QtCore.QRegExp("[0-9]+")
        validator = QtGui.QValidator = QtGui.QRegExpValidator(rx)
        self.ui.lineEdit_4.setValidator(validator)

        rx = QtCore.QRegExp("[а-яА-Я]+")
        validator = QtGui.QValidator = QtGui.QRegExpValidator(rx)
        self.ui.lineEdit_2.setValidator(validator)

        rx = QtCore.QRegExp("[а-яА-Я]+")
        validator = QtGui.QValidator = QtGui.QRegExpValidator(rx)
        self.ui.lineEdit_6.setValidator(validator)

        rx = QtCore.QRegExp("[а-яА-Я]+")
        validator = QtGui.QValidator = QtGui.QRegExpValidator(rx)
        self.ui.lineEdit_7.setValidator(validator)

        rx = QtCore.QRegExp("[а-яА-Я]+")
        validator = QtGui.QValidator = QtGui.QRegExpValidator(rx)
        self.ui.lineEdit_9.setValidator(validator)

        rx = QtCore.QRegExp("[0-9]+")
        validator = QtGui.QValidator = QtGui.QRegExpValidator(rx)
        self.ui.lineEdit_10.setValidator(validator)

        
        self.ui.lineEdit_15.setInputMask("0000-00-00")
        self.ui.label_26.setText('Формат даты: гггг-мм-дд')

        self.fu = 0

        

    def Pods(self):
        # Вспоывающие подсказки запросов
        self.ui.listWidget.item(0).setToolTip('Вывести стоимость определенной услуги.')
        self.ui.listWidget.item(1).setToolTip('Сколько существует специальностей?')
        self.ui.listWidget.item(2).setToolTip('Вывести ФИО рабочих, которые работали\nв определённую дату.')
        self.ui.listWidget.item(3).setToolTip('Вывести ФИО клиента, делавшего заказ\nв определённую дату.')
        self.ui.listWidget.item(4).setToolTip('Какие услуги были выполнены в определенный день?')
        self.ui.listWidget.item(5).setToolTip('Вывести наименование самой дорогой услуги.')
        self.ui.listWidget.item(6).setToolTip('Вывести наименование самой дешевой услуги.')
        self.ui.listWidget.item(7).setToolTip('Вывести ФИО клиентов по id_регистрации.')
        self.ui.listWidget.item(8).setToolTip('Вывести ФИО и квалификацию рабочего\nпо id_регистрации.')
        self.ui.listWidget.item(9).setToolTip('Вывести id_регистрации, ФИО клиента,\nФИО рабочего и наименование услуги.')


    def pri(self): # Функция блокировки запросов в других таблицах
        numberTab = self.ui.tabWidget.currentIndex()
        if self.ui.tabWidget.currentIndex() == 0:
            self.ui.stackedWidget_2.setCurrentIndex(1)
            self.ui.label_24.setText('')
            self.ui.label_25.setText('')
        elif self.ui.tabWidget.currentIndex() == 1:
            self.ui.stackedWidget_2.setCurrentIndex(1)
            self.ui.label_24.setText('')
            self.ui.label_25.setText('')
        elif self.ui.tabWidget.currentIndex() == 2:
            self.ui.stackedWidget_2.setCurrentIndex(1)
            self.ui.label_24.setText('')
            self.ui.label_25.setText('')
        elif self.ui.tabWidget.currentIndex() == 3:
            self.ui.stackedWidget_2.setCurrentIndex(1)
            self.ui.label_24.setText('')
            self.ui.label_25.setText('')
        else:
            self.ui.stackedWidget_2.setCurrentIndex(0)
            self.ui.label_24.setText('')
            self.ui.label_25.setText('')


    def Instruk(self): # Инструкция
        os.startfile('Instr.txt')


    def Seal(self, x): # Сохранение
        numberTab = self.ui.tabWidget.currentIndex()

        wb = openpyxl.Workbook()
        wb.create_sheet(title = 'Первый лист', index = 0)
        sheet = wb['Первый лист']

        if numberTab == 0:
            try:
                connection = sqlite3.connect('registr.db')
                cursor = connection.cursor()
                cursor.execute('SELECT * FROM Клиенты')
                results = cursor.fetchall()
                k = results
                tble = []
                if x == 1:
                    my_file = open('Клиенты.doc', 'w')
                    for j in range(len(k)):
                        kuk = 4
                        for i in range(len(k[0])):
                            tb = self.ui.tableWidget_1.item(j, i).text()
                            my_file.write(tb+' ')
                            kuk = kuk - 1
                            if kuk == 0:
                                my_file.write('\n')
                    QMessageBox.about(self, "Успешно", "Данные таблицы сохранены в файле 'Клиенты.doc'")
                    my_file.close()

                    
                if x == 2:
                    my_file = open('Клиенты.txt', 'w')
                    for j in range(len(k)):
                        kuk = 4
                        for i in range(len(k[0])):
                            tb = self.ui.tableWidget_1.item(j, i).text()
                            my_file.write(tb+' ')
                            kuk = kuk - 1
                            if kuk == 0:
                                my_file.write('\n')
                    QMessageBox.about(self, "Успешно", "Данные таблицы сохранены в файле 'Клиенты.txt'")
                    my_file.close()

                if x == 3:
                    for j in range(1, len(k)+1):
                        for i in range(1, len(k[0])+1):
                            tb = self.ui.tableWidget_1.item(j-1, i-1).text()
                            cell = sheet.cell(row = j, column = i)
                            cell.value = tb

                    wb.save('Клиенты.xlsx')
                    QMessageBox.about(self, "Успешно", "Данные таблицы сохранены в файле 'Клиенты.xlsx'")
                                                                                 

                connection.commit()
                connection.close()
             
                
            except:
                QMessageBox.about(self, "Невозможно", "Отсутствуют записи в таблице!")
                
        if numberTab == 1:
            try:
                connection = sqlite3.connect('registr.db')
                cursor = connection.cursor()
                cursor.execute('SELECT * FROM Работники')
                results = cursor.fetchall()
                k = results
                tble = []
                if x == 1:
                    my_file = open('Работники.doc', 'w')
                    for j in range(len(k)):
                        kuk = 3
                        for i in range(len(k[0])):
                            tb = self.ui.tableWidget_2.item(j, i).text()
                            my_file.write(tb+' ')
                            kuk = kuk - 1
                            if kuk == 0:
                                my_file.write('\n')
                    QMessageBox.about(self, "Успешно", "Данные таблицы сохранены в файле 'Работники.doc'")
                    my_file.close()

                if x == 2:
                    my_file = open('Работники.txt', 'w')
                    for j in range(len(k)):
                        kuk = 3
                        for i in range(len(k[0])):
                            tb = self.ui.tableWidget_2.item(j, i).text()
                            my_file.write(tb+' ')
                            kuk = kuk - 1
                            if kuk == 0:
                                my_file.write('\n')
                    QMessageBox.about(self, "Успешно", "Данные таблицы сохранены в файле 'Работники.txt'")
                    my_file.close()


                if x == 3:
                    for j in range(1, len(k)):
                        for i in range(1, len(k[0])):
                            tb = self.ui.tableWidget_2.item(j-1, i-1).text()
                            cell = sheet.cell(row = j, column = i)
                            cell.value = tb

                    wb.save('Работники.xlsx')
                    QMessageBox.about(self, "Успешно", "Данные таблицы сохранены в файле 'Работники.xlsx'")

                connection.commit()
                connection.close()
             
                
            except:
                QMessageBox.about(self, "Невозможно", "Отсутствуют записи в таблице!")
            

        if numberTab == 2:
            try:
                connection = sqlite3.connect('registr.db')
                cursor = connection.cursor()
                cursor.execute('SELECT * FROM Услуги')
                results = cursor.fetchall()
                k = results
                tble = []
                if x == 1:
                    my_file = open('Услуги.doc', 'w')
                    for j in range(len(k)):
                        kuk = 3
                        for i in range(len(k[0])):
                            tb = self.ui.tableWidget_3.item(j, i).text()
                            my_file.write(tb+' ')
                            kuk = kuk - 1
                            if kuk == 0:
                                my_file.write('\n')
                    QMessageBox.about(self, "Успешно", "Данные таблицы сохранены в файле 'Услуги.doc'")
                    my_file.close()

                if x == 2:
                    my_file = open('Услуги.txt', 'w')
                    for j in range(len(k)):
                        kuk = 3
                        for i in range(len(k[0])):
                            tb = self.ui.tableWidget_3.item(j, i).text()
                            my_file.write(tb+' ')
                            kuk = kuk - 1
                            if kuk == 0:
                                my_file.write('\n')
                    QMessageBox.about(self, "Успешно", "Данные таблицы сохранены в файле 'Услуги.txt'")
                    my_file.close()

                if x == 3:
                    for j in range(1, len(k)):
                        for i in range(1, len(k[0])):
                            tb = self.ui.tableWidget_3.item(j-1, i-1).text()
                            cell = sheet.cell(row = j, column = i)
                            cell.value = tb

                    wb.save('Услуги.xlsx')
                    QMessageBox.about(self, "Успешно", "Данные таблицы сохранены в файле 'Услуги.xlsx'")

                connection.commit()
                connection.close()
             
                
            except:
                QMessageBox.about(self, "Невозможно", "Отсутствуют записи в таблице!")

        
        if numberTab == 3:
            try:
                connection = sqlite3.connect('registr.db')
                cursor = connection.cursor()
                cursor.execute('SELECT * FROM Регистрация_работ')
                results = cursor.fetchall()
                k = results
                tble = []
                if x == 1:
                    my_file = open('Регистрация_работ.doc', 'w')
                    for j in range(len(k)):
                        kuk = 5
                        for i in range(len(k[0])):
                            tb = self.ui.tableWidget_4.item(j, i).text()
                            my_file.write(tb+' ')
                            kuk = kuk - 1
                            if kuk == 0:
                                my_file.write('\n')
                    QMessageBox.about(self, "Успешно", "Данные таблицы сохранены в файле 'Регистрация_работ.doc'")
                    my_file.close()

                if x == 2:
                    my_file = open('Регистрация_работ.txt', 'w')
                    for j in range(len(k)):
                        kuk = 5
                        for i in range(len(k[0])):
                            tb = self.ui.tableWidget_4.item(j, i).text()
                            my_file.write(tb+' ')
                            kuk = kuk - 1
                            if kuk == 0:
                                my_file.write('\n')
                    QMessageBox.about(self, "Успешно", "Данные таблицы сохранены в файле 'Регистрация_работ.txt'")
                    my_file.close()

                if x == 3:
                    for j in range(1, len(k)):
                        for i in range(1, len(k[0])):
                            tb = self.ui.tableWidget_4.item(j-1, i-1).text()
                            cell = sheet.cell(row = j, column = i)
                            cell.value = tb

                    wb.save('Регистрация_работ.xlsx')
                    QMessageBox.about(self, "Успешно", "Данные таблицы сохранены в файле 'Регистрация_работ.xlsx'")

                connection.commit()
                connection.close()
             
                
            except:
                QMessageBox.about(self, "Невозможно", "Отсутствуют записи в таблице!")


    def Prog(self): # О программе
        QMessageBox.about(self, "О программе", "Программа разработана курсантом 431 группы\nАвтор: Шабалин Дмитрий Алексеевич\n\nПрограмма работает с базой данных 'registr'. Для того чтобы добавить, изменить или удалить записи следует нажать соответствующие кнопки. Программа расчитана на два вида пользователей: Администратор и Пользователь. Чтобы переключиться между режимами необходимо нажать соответствующий пункт в меню.")

    def Admin(self): # Переход в режим администратора
        fur = True
        if self.flacon == 1 :
            QMessageBox.about(self, "Невозможно", "Вы уже в режиме администратора!")
        else:
            if self.ing == '0':
                while fur == True:
                    text, ok = QtWidgets.QInputDialog.getText(self, 'Регистрация', 'Придумайте пароль:')
                    if ok:
                        if text == '0':
                            QMessageBox.critical(self, "Невозможно", "Пароль слишком легкий!")
                        elif text == '':
                            QMessageBox.critical(self, "Невозможно", "Необходимо ввести пароль")
                        else:
                            my_file = open('Pass.txt', 'w')
                            my_file.write(text)
                            my_file.close()
                            fur = False
                    fur = False
                    
                   
                text, ok = QtWidgets.QInputDialog.getText(self, 'Проверка', 'Введите пароль:')       
                if ok:
                    f = open('Pass.txt','r')
                    for i in f:
                        pas = i
                        if text == i:
                            QMessageBox.about(self, "Успешно", "Вы перешли в режим администратора!")

                            self.ui.pushButton_2.setEnabled(True)
                            self.ui.pushButton_3.setEnabled(True)

                            self.ui.pushButton_2.setStyleSheet('''QPushButton {
                                                             background-color: grey;
                                                                color: white;
                                                         }''')

                            self.ui.pushButton_3.setStyleSheet('''QPushButton {
                                                             background-color: grey;
                                                                color: white;
                                                         }''')

                            self.flacon = 1

                            self.ing = i
                        else:
                            QMessageBox.critical(self, "Ошибка", "Некорректные данные!")


            else:
                text, ok = QtWidgets.QInputDialog.getText(self, 'Проверка', 'Введите пароль:')       
                if ok:
                    f = open('Pass.txt','r')
                    for i in f:
                        pas = i
                        if text == i:
                            QMessageBox.about(self, "Успешно", "Вы перешли в режим администратора!")
                            self.ui.pushButton_2.setEnabled(True)
                            self.ui.pushButton_3.setEnabled(True)

                            self.ui.pushButton_2.setStyleSheet('''QPushButton {
                                                             background-color: grey;
                                                                color: white;
                                                         }''')

                            self.ui.pushButton_3.setStyleSheet('''QPushButton {
                                                             background-color: grey;
                                                                color: white;
                                                         }''')

                            self.flacon = 1
                            
                        else:
                            QMessageBox.critical(self, "Ошибка", "Некорректные данные!")
                                           


    def Pol(self): # Переход в режим пользователя
        if self.flacon == 0:
            QMessageBox.about(self, "Невозможно", "Вы уже в режиме пользователя!")
        else:
            self.flacon = 0
            QMessageBox.about(self, "Успешно", "Вы перешли в режим пользователя!")

            
            self.ui.pushButton_2.setEnabled(False)
            self.ui.pushButton_3.setEnabled(False)

            self.ui.pushButton_2.setStyleSheet('''QPushButton {
                                            background-color: ##424242;
                                            color: grey;
                                            }''')

            self.ui.pushButton_3.setStyleSheet('''QPushButton {
                                            background-color: ##424242;
                                            color: grey;
                                            }''')            

    def Cancel(self): # Функция кнопок отмены
        self.ui.stackedWidget.setCurrentIndex(0)
        
    def Podkl(self): # Подключение к БД и вывод данных
        try:
        
            connection = sqlite3.connect('registr.db')
            cursor = connection.cursor()
            
            numberTab = self.ui.tabWidget.currentIndex()
            
            if numberTab == 0:
                cursor.execute('SELECT * FROM Клиенты')
                self.pre = 1
                self.ui.statusbar.showMessage("База данных 'registr', таблица 'Клиенты'")
            if numberTab == 1:
                cursor.execute('SELECT * FROM Работники')
                self.ui.statusbar.showMessage("База данных 'registr', таблица 'Работники'")
            if numberTab == 2:
                cursor.execute('SELECT * FROM Услуги')
                self.ui.statusbar.showMessage("База данных 'registr', таблица 'Услуги'")
                
            if numberTab == 3:
                cursor.execute('SELECT * FROM Регистрация_работ')
                self.ui.statusbar.showMessage("База данных 'registr', таблица 'Регистрация_работ")
            if numberTab == 4:
                numberlist = self.ui.listWidget.currentRow()
                if numberlist == 0:
                    self.ui.statusbar.showMessage("Вывести стоимость определенной услуги.")
                    text, ok = QtWidgets.QInputDialog.getText(self, 'Уточним..', 'Введите название услуги:')       
                    if ok:
                        cursor.execute('SELECT наименование, стоимость FROM Услуги WHERE наименование = ' + '"' + text + '"')
                        
            
                if numberlist == 1:
                    cursor.execute('select квалификация from Работники')
                    self.ui.statusbar.showMessage("Сколько существует специальностей?")
                    
                if numberlist == 2:
                    self.ui.statusbar.showMessage("Вывести ФИО рабочих, которые работали в определённую дату.")
                    text, ok = QtWidgets.QInputDialog.getText(self, 'Уточним..', 'Формат даты: гггг-мм-дд\nВведите дату:')       
                    if ok:
                        cursor.execute('select ФИО, дата from Работники, Регистрация_работ where Работники.id_рабочего = Регистрация_работ.id_рабочего and дата = '+ '"' + text + '"')
                        

                if numberlist == 3:
                    self.ui.statusbar.showMessage("Вывести ФИО клиента, делавшего заказ в определённую дату.")
                    text, ok = QtWidgets.QInputDialog.getText(self, 'Уточним..', 'Формат даты: гггг-мм-дд\nВведите дату:')       
                    if ok:
                        cursor.execute('select ФИО, дата from Клиенты, Регистрация_работ where Клиенты.id_клиента = Регистрация_работ.id_клиента and дата = '+ '"' + text + '"')
                        

                if numberlist == 4:
                    self.ui.statusbar.showMessage("Какие услуги были выполнены в определенный день?")
                    text, ok = QtWidgets.QInputDialog.getText(self, 'Уточним..', 'Формат даты: гггг-мм-дд\nВведите дату:')       
                    if ok:
                        cursor.execute('select наименование, дата from Услуги, Регистрация_работ where Услуги.id_услуги = Регистрация_работ.id_услуги and дата = '+ '"' + text + '"')
                        
                    
                if numberlist == 5:
                    cursor.execute('select наименование, стоимость from Услуги group by стоимость order by стоимость desc limit 1')
                    self.ui.statusbar.showMessage("Вывести наименование самой дорогой услуги.")

                if numberlist == 6:
                    cursor.execute('select наименование, стоимость from Услуги group by стоимость order by стоимость asc limit 1')
                    self.ui.statusbar.showMessage("Вывести наименование самой дешевой услуги.")

                if numberlist == 7:
                    self.ui.statusbar.showMessage("Вывести ФИО клиентов по id_регистрации.")
                    text, ok = QtWidgets.QInputDialog.getText(self, 'Уточним..', 'Введите id_регистрации:')       
                    if ok:
                        cursor.execute('select ФИО, id_регистрации from Клиенты, Регистрация_работ where Клиенты.id_клиента = Регистрация_работ.id_клиента and Регистрация_работ.id_регистрации ='+ '"' + text + '"')
                        

                if numberlist == 8:
                    self.ui.statusbar.showMessage("Вывести ФИО и квалификацию рабочего по id_регистрации.")
                    text, ok = QtWidgets.QInputDialog.getText(self, 'Уточним..', 'Введите id_регистрации:')       
                    if ok:
                        cursor.execute('select ФИО, квалификация, id_регистрации from Работники, Регистрация_работ where Работники.id_рабочего = Регистрация_работ.id_рабочего and Регистрация_работ.id_регистрации ='+ '"' + text + '"')
                        

                if numberlist == 9:
                    cursor.execute('select id_регистрации, Клиенты.ФИО, Работники.ФИО, Услуги.наименование from Регистрация_работ, Клиенты, Работники, Услуги where Клиенты.id_клиента = Регистрация_работ.id_клиента and Работники.id_рабочего = Регистрация_работ.id_рабочего and Услуги.id_услуги = Регистрация_работ.id_услуги')
                    self.ui.statusbar.showMessage("Вывести id_регистрации, ФИО клиента, ФИО рабочего и наименование услуги.")

                
            results = cursor.fetchall()
            k = results

            table = eval('self.ui.tableWidget_' + str(int(numberTab)+1))
            
            self.insTb(k, table)
            connection.commit()
            connection.close()

            
        except:
            pass
        
    def insTb(self,k,tbl):  #Функция заполнения таблицы

        sch = 0
        tbl.setRowCount(len(k))
        tbl.setColumnCount(len(k[0]))
        for j in range(len(k)):
            sch+=1
            for i in range(len(k[0])):
                tbl.setItem(j, i, QWT(str(k[j][i])))

        if self.fu == 0:
                
            for c in range(self.ui.tableWidget_1.rowCount()):
                idid = self.ui.tableWidget_1.item(c, 0).text()
                self.ui.comboBox.addItem(idid)
        else:
            pass
        if self.pre == 1:
            self.fu = 1


        # Форматирование по размеру
        self.ui.tableWidget_1.resizeColumnsToContents()
        self.ui.tableWidget_1.resizeRowsToContents()
        self.ui.tableWidget_2.resizeColumnsToContents()
        self.ui.tableWidget_2.resizeRowsToContents()
        self.ui.tableWidget_3.resizeColumnsToContents()
        self.ui.tableWidget_3.resizeRowsToContents()
        self.ui.tableWidget_4.resizeColumnsToContents()
        self.ui.tableWidget_4.resizeRowsToContents()
        self.ui.tableWidget_5.resizeColumnsToContents()
        self.ui.tableWidget_5.resizeRowsToContents()

        self.ui.tableWidget_1.verticalHeader().hide()
        self.ui.tableWidget_2.verticalHeader().hide()
        self.ui.tableWidget_3.verticalHeader().hide()
        self.ui.tableWidget_4.verticalHeader().hide()


        
        self.ui.label_24.setText('Количество записей:')
        self.ui.label_25.setText(str(sch))
        self.ui.label_25.setGeometry(QtCore.QRect(120, 380, 131, 16))
        

        numberTab = self.ui.tabWidget.currentIndex()
        if numberTab == 2:
            x = 0
            for i in range(sch):
                itog = self.ui.tableWidget_3.item(i, 2).text()
                x = x + int(itog)
            self.ui.label_24.setText('Итого:')
            self.ui.label_25.setText(str(x))
            self.ui.label_25.setGeometry(QtCore.QRect(45, 380, 131, 16))

    def Update(self): # Проверка текущей таблицы
        numberPage = self.ui.stackedWidget.currentIndex()

        numberTab = self.ui.tabWidget.currentIndex()
        
        if numberTab == 0:
            self.ui.stackedWidget.setCurrentIndex(1)
        if numberTab == 1:
            self.ui.stackedWidget.setCurrentIndex(2)
        if numberTab == 2:
            self.ui.stackedWidget.setCurrentIndex(3)
        if numberTab == 3:
            QMessageBox.about(self, "Предупреждение", "Если таблица 'Клиенты' не обновлена, пожалуйста, обновите!")
            self.ui.stackedWidget.setCurrentIndex(4)

    def upd1(self): # Изменение записей
        try:
            connection = sqlite3.connect('registr.db')
            cursor = connection.cursor()
            
            a = []
            a.append(self.ui.lineEdit.text())
            a.append(self.ui.lineEdit_2.text())
            a.append(self.ui.lineEdit_3.text())
            a.append(self.ui.lineEdit_4.text())
            a=tuple(a)

            numberRow = int(self.ui.spinBox.text())
            idd = (self.ui.tableWidget_1.item(numberRow - 1, 0).text())
            
            cursor.execute('UPDATE Клиенты SET id_клиента = ?, ФИО = ?, адрес = ?, телефон = ? WHERE id_клиента=' + str(idd),a)

            connection.commit()
            connection.close()
            self.Podkl()
            
            self.ui.stackedWidget.setCurrentIndex(0)
        except:
            QMessageBox.critical(self, "Ошибка", "Введенные данные некорректны!")


    def upd2(self): # Изменение записей
        try:
            connection = sqlite3.connect('registr.db')
            cursor = connection.cursor()
            a = []
            a.append(self.ui.lineEdit_5.text())
            a.append(self.ui.lineEdit_6.text())
            a.append(self.ui.lineEdit_7.text())
            a=tuple(a)

            numberRow = int(self.ui.spinBox_2.text())
            idd = (self.ui.tableWidget_2.item(numberRow - 1, 0).text())
            
            cursor.execute('UPDATE Работники SET id_рабочего = ?, ФИО = ?, квалификация = ? WHERE id_рабочего=' + str(idd),a)

            connection.commit()
            connection.close()
            self.Podkl()
            
            self.ui.stackedWidget.setCurrentIndex(0)
        except:
            QMessageBox.critical(self, "Ошибка", "Введенные данные некорректны!")


    def upd3(self): # Изменение записей
        try:
            connection = sqlite3.connect('registr.db')
            cursor = connection.cursor()
            
            a = []
            a.append(self.ui.lineEdit_8.text())
            a.append(self.ui.lineEdit_9.text())
            a.append(self.ui.lineEdit_10.text())
            a=tuple(a)

            numberRow = int(self.ui.spinBox_3.text())
            idd = (self.ui.tableWidget_3.item(numberRow - 1, 0).text())
            
            cursor.execute('UPDATE Услуги SET id_услуги = ?, наименование = ?, стоимость = ? WHERE id_услуги=' + str(idd),a)

            connection.commit()
            connection.close()
            self.Podkl()
            
            self.ui.stackedWidget.setCurrentIndex(0)
        except:
            QMessageBox.critical(self, "Ошибка", "Введенные данные некорректны!")

    def upd4(self): # Изменение записей
        try:
            connection = sqlite3.connect('registr.db')
            cursor = connection.cursor()
            
            a = []
            a.append(self.ui.lineEdit_11.text())
            #a.append(self.ui.lineEdit_12.text())
            a.append(self.ui.comboBox.currentText())
            #a.append(self.ui.lineEdit_13.text())
            a.append(self.ui.comboBox_2.currentText())
            #a.append(self.ui.lineEdit_14.text())
            a.append(self.ui.comboBox_3.currentText())
            a.append(self.ui.lineEdit_15.text())
            a=tuple(a)

            numberRow = int(self.ui.spinBox_4.text())
            idd = (self.ui.tableWidget_4.item(numberRow - 1, 0).text())
            
            cursor.execute('UPDATE Регистрация_работ SET id_регистрации = ?, id_клиента = ?, id_рабочего = ?, id_услуги = ?, дата = ? WHERE id_регистрации =' + str(idd),a)

            connection.commit()
            connection.close()
            self.Podkl()
            
            self.ui.stackedWidget.setCurrentIndex(0)
        except:
            QMessageBox.critical(self, "Ошибка", "Введенные данные некорректны!")


    def add1(self): # Добавление записей
        try:
            connection = sqlite3.connect('registr.db')
            cursor = connection.cursor()
            a = []

            #self.ui.lineEdit_4.setInputMask("AAAAA")
            
            a.append(self.ui.lineEdit.text())
            a.append(self.ui.lineEdit_2.text())
            a.append(self.ui.lineEdit_3.text())
            a.append(self.ui.lineEdit_4.text())
            
            a=tuple(a)
                
            
            cursor.execute('INSERT INTO Клиенты VALUES(?,?,?,?)',a)

            connection.commit()
            connection.close()
            self.Podkl()

            self.ui.stackedWidget.setCurrentIndex(0)
        except:
            QMessageBox.critical(self, "Ошибка", "Введенные данные некорректны!")

    def add2(self): # Добавление записей
        try:
            connection = sqlite3.connect('registr.db')
            cursor = connection.cursor()
            a = []
            a.append(self.ui.lineEdit_5.text())
            a.append(self.ui.lineEdit_6.text())
            a.append(self.ui.lineEdit_7.text())
            a=tuple(a)

            cursor.execute('INSERT INTO Работники VALUES(?,?,?)',a)

            connection.commit()
            connection.close()
            self.Podkl()

            self.ui.stackedWidget.setCurrentIndex(0)
        except:
            QMessageBox.critical(self, "Ошибка", "Введенные данные некорректны!")

    def add3(self): # Добавление записей
        try:
            connection = sqlite3.connect('registr.db')
            cursor = connection.cursor()
            a = []
            a.append(self.ui.lineEdit_8.text())
            a.append(self.ui.lineEdit_9.text())
            a.append(self.ui.lineEdit_10.text())
            a=tuple(a)

            cursor.execute('INSERT INTO Услуги VALUES(?,?,?)',a)

            connection.commit()
            connection.close()
            self.Podkl()

            self.ui.stackedWidget.setCurrentIndex(0)
        except:
            QMessageBox.critical(self, "Ошибка", "Введенные данные некорректны!")

    def add4(self): # Добавление записей
        try:
            connection = sqlite3.connect('registr.db')
            cursor = connection.cursor()
            a = []
            a.append(self.ui.lineEdit_11.text())
            #a.append(self.ui.lineEdit_12.text())
            a.append(self.ui.comboBox.currentText())
            #a.append(self.ui.lineEdit_13.text())
            a.append(self.ui.comboBox_2.currentText())
            #a.append(self.ui.lineEdit_14.text())
            a.append(self.ui.comboBox_3.currentText())
            a.append(self.ui.lineEdit_15.text())
            a=tuple(a)

            cursor.execute('INSERT INTO Регистрация_работ VALUES(?,?,?,?,?)',a)

            connection.commit()
            connection.close()
            self.Podkl()

            self.ui.stackedWidget.setCurrentIndex(0)
        except:
            QMessageBox.critical(self, "Ошибка", "Введенные данные некорректны!")


    def del1(self): # Удаление записей
        try:
            numberTab = self.ui.tabWidget.currentIndex()
            if numberTab == 4:
                pass
            else:
                text, ok = QtWidgets.QInputDialog.getText(self, 'Удаление', 'Введите номер строки:')
                if ok:
                    connection = sqlite3.connect('registr.db')
                    cursor = connection.cursor()
                    
                    numberTab = self.ui.tabWidget.currentIndex()

                    res = text
                    
                    if numberTab == 0:
                        cursor.execute('DELETE FROM Клиенты WHERE id_клиента=?', (self.ui.tableWidget_1.item(int(res) - 1, 0).text(),))
                    if numberTab == 1:
                        cursor.execute('DELETE FROM Работники WHERE id_рабочего=?', (self.ui.tableWidget_2.item(int(res) - 1, 0).text(),))
                    if numberTab == 2:
                        cursor.execute('DELETE FROM Услуги WHERE id_услуги=?', (self.ui.tableWidget_3.item(int(res) - 1, 0).text(),))
                    if numberTab == 3:
                        cursor.execute('DELETE FROM Регистрация_работ WHERE id_регистрации=?', (self.ui.tableWidget_4.item(int(res) - 1, 0).text(),))
                        
                    connection.commit()
                    connection.close()
                    self.Podkl()

                    self.ui.stackedWidget.setCurrentIndex(0)

        except:
            QMessageBox.critical(self, "Ошибка", "Данная запись отсутствует!")


        
if __name__=="__main__":
    app = QtWidgets.QApplication(sys.argv)
    myapp = MyWin()
    myapp.show()
    sys.exit(app.exec_())

